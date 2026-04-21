"""
Excel Extractor Agent - Specialized in Excel file extraction
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional, Union
import structlog
from pathlib import Path

from .base_extractor import BaseExtractorAgent

logger = structlog.get_logger()


class ExcelExtractorAgent(BaseExtractorAgent):
    """
    Specialized agent for Excel extraction

    Capabilities:
    - Read XLSX, XLS files
    - Extract specific sheets
    - Extract ranges
    - Read formulas
    - Extract cell formatting
    - Handle merged cells
    - Multiple sheets to single DataFrame
    - Excel to JSON/CSV conversion
    """

    def _define_capabilities(self) -> List[str]:
        return [
            "read_excel",
            "extract_sheet",
            "extract_range",
            "read_formulas",
            "extract_formatting",
            "handle_merged_cells",
            "multi_sheet_extraction",
            "excel_to_csv",
            "excel_to_json"
        ]

    def _get_supported_formats(self) -> List[str]:
        return [".xlsx", ".xls", ".xlsm", ".xlsb"]

    async def validate(self, source: Any) -> bool:
        """Validate if source is valid Excel"""
        try:
            if isinstance(source, (str, Path)):
                pd.read_excel(source, nrows=1)
            return True
        except Exception as e:
            logger.error("Excel validation failed", error=str(e))
            return False

    async def extract(self, source: Union[str, Path],
                     query: Optional[Dict[str, Any]] = None) -> Any:
        """
        Extract data from Excel

        Args:
            source: Excel file path
            query: {
                'sheet_name': 'Sheet1' | 0 | ['Sheet1', 'Sheet2'],
                'header_row': 0,  # Which row contains headers
                'skip_rows': 0,  # Rows to skip
                'use_cols': 'A:D' | [0, 1, 2],  # Columns to read
                'range': 'A1:D10',  # Specific range
                'read_formulas': True | False,
                'output_format': 'dict' | 'dataframe' | 'json'
            }

        Returns:
            Extracted data
        """
        query = query or {}

        sheet_name = query.get('sheet_name', 0)
        header_row = query.get('header_row', 0)
        skip_rows = query.get('skip_rows', 0)
        use_cols = query.get('use_cols', None)
        output_format = query.get('output_format', 'dict')

        # Read Excel
        if query.get('range'):
            # Read specific range
            df = self._read_range(source, query['range'], sheet_name)
        else:
            # Normal read
            df = pd.read_excel(
                source,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows,
                usecols=use_cols
            )

        # Handle multiple sheets
        if isinstance(df, dict):
            # Multiple sheets returned as dict
            result = {}
            for name, sheet_df in df.items():
                if output_format == 'dict':
                    result[name] = sheet_df.to_dict(orient='records')
                elif output_format == 'json':
                    result[name] = sheet_df.to_json(orient='records')
                else:
                    result[name] = sheet_df
            return result

        # Single sheet
        if output_format == 'dict':
            return df.to_dict(orient='records')
        elif output_format == 'json':
            return df.to_json(orient='records')
        elif output_format == 'dataframe':
            return df
        else:
            return df.to_dict(orient='records')

    def _read_range(self, source: Union[str, Path],
                   cell_range: str,
                   sheet_name: Union[str, int] = 0) -> pd.DataFrame:
        """Read specific range from Excel"""
        wb = openpyxl.load_workbook(source, data_only=True)

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        # Parse range and extract data
        data = []
        for row in ws[cell_range]:
            data.append([cell.value for cell in row])

        # Convert to DataFrame
        if len(data) > 0:
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame()

        return df

    async def get_sheet_names(self, source: Union[str, Path]) -> List[str]:
        """Get list of sheet names"""
        excel_file = pd.ExcelFile(source)
        return excel_file.sheet_names

    async def extract_sheet(self, source: Union[str, Path],
                           sheet_name: Union[str, int]) -> List[Dict[str, Any]]:
        """Extract specific sheet"""
        return await self.extract(source, {
            'sheet_name': sheet_name,
            'output_format': 'dict'
        })

    async def extract_all_sheets(self, source: Union[str, Path]) -> Dict[str, List[Dict]]:
        """Extract all sheets"""
        return await self.extract(source, {
            'sheet_name': None,  # None = all sheets
            'output_format': 'dict'
        })

    async def extract_with_formulas(self, source: Union[str, Path],
                                   sheet_name: Union[str, int] = 0) -> pd.DataFrame:
        """Extract data with formulas preserved"""
        wb = openpyxl.load_workbook(source)

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        data = []
        for row in ws.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    row_data.append({
                        'value': cell.value,
                        'formula': f'={cell.value}'
                    })
                else:
                    row_data.append(cell.value)
            data.append(row_data)

        return pd.DataFrame(data[1:], columns=data[0])

    async def get_cell_formatting(self, source: Union[str, Path],
                                  sheet_name: Union[str, int] = 0,
                                  cell: str = 'A1') -> Dict[str, Any]:
        """Get cell formatting information"""
        wb = openpyxl.load_workbook(source)

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        cell_obj = ws[cell]

        return {
            'value': cell_obj.value,
            'font': {
                'name': cell_obj.font.name,
                'size': cell_obj.font.size,
                'bold': cell_obj.font.bold,
                'italic': cell_obj.font.italic,
                'color': cell_obj.font.color.rgb if cell_obj.font.color else None
            },
            'fill': {
                'pattern': cell_obj.fill.patternType,
                'color': cell_obj.fill.fgColor.rgb if cell_obj.fill.fgColor else None
            },
            'alignment': {
                'horizontal': cell_obj.alignment.horizontal,
                'vertical': cell_obj.alignment.vertical
            },
            'number_format': cell_obj.number_format
        }

    async def excel_to_csv(self, source: Union[str, Path],
                          output_path: str,
                          sheet_name: Union[str, int] = 0):
        """Convert Excel to CSV"""
        df = pd.read_excel(source, sheet_name=sheet_name)
        df.to_csv(output_path, index=False)
        logger.info(f"Converted {source} to {output_path}")

    async def excel_to_json(self, source: Union[str, Path],
                           output_path: str,
                           sheet_name: Union[str, int] = 0):
        """Convert Excel to JSON"""
        df = pd.read_excel(source, sheet_name=sheet_name)
        df.to_json(output_path, orient='records', indent=2)
        logger.info(f"Converted {source} to {output_path}")

    async def get_metadata(self, source: Union[str, Path]) -> Dict[str, Any]:
        """Get Excel file metadata"""
        wb = openpyxl.load_workbook(source)

        return {
            'num_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'properties': {
                'creator': wb.properties.creator,
                'last_modified_by': wb.properties.lastModifiedBy,
                'created': str(wb.properties.created),
                'modified': str(wb.properties.modified),
                'title': wb.properties.title,
                'subject': wb.properties.subject
            }
        }

    async def search_value(self, source: Union[str, Path],
                          search_term: str,
                          sheet_name: Union[str, int] = 0) -> List[Dict[str, Any]]:
        """Search for value in Excel file"""
        wb = openpyxl.load_workbook(source)

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        results = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and search_term in str(cell.value):
                    results.append({
                        'cell': cell.coordinate,
                        'row': cell.row,
                        'column': cell.column,
                        'value': cell.value
                    })

        return results
